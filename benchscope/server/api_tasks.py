"""任务管理 API：创建、启动、停止、删除、查询任务。"""
from __future__ import annotations

import logging

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel

from benchscope.server.state import state

log = logging.getLogger("benchscope.api_tasks")

router = APIRouter(prefix="/api/tasks", tags=["tasks"])

MAX_LOG_LINES = 8000


class CreateTaskRequest(BaseModel):
    framework: str = "vllm"
    model: str = ""
    tokenizer: str = ""
    dataset: dict = {}
    concurrency_list: list = []
    gpu: dict = {}
    request_rate: str | float = "inf"
    tpot_threshold_ms: float | None = None
    # 性能测试模式：concurrency（并发模式）/ threshold（阈值模式）
    mode: str = "concurrency"
    # 阈值模式：TTFT / TPOT 阈值（ms）与统计量（mean/median/p99），0 表示该指标不参与判定
    ttft_threshold_ms: float = 0
    ttft_statistic: str = "mean"
    tpot_statistic: str = "mean"
    # 阈值模式：Output token throughput 上限（tok/s），0 表示不参与核心逻辑
    output_throughput_threshold: float = 0
    # 阈值模式：单组最大并发搜索上限（默认 4096）
    max_concurrency_search: int = 4096
    # Step2「性能参数」编辑后的框架默认参数 yaml 内容（跟随进入命令）
    params_yaml: dict = {}


class UpdateThresholdRequest(BaseModel):
    tpot_threshold_ms: float
    precision: str = ""
    curated: dict = {}
    extra_args: list = []


class ExportExcelRequest(BaseModel):
    """前端当前表格内容导出：headers 为列标题（与 values 顺序对齐），rows 为数据行。"""
    headers: list[str] = []
    rows: list[dict] = []  # 每项 {"values": [...], "group": bool}；group 行作为组标题加粗显示


@router.get("")
def list_tasks():
    return {"tasks": state.tasks.list_tasks()}


@router.post("/preview")
def preview_new_task(req: CreateTaskRequest):
    """预览将要执行的命令（创建任务前调用，无需 task_id）。"""
    from benchscope.server.test_manager import build_command_lines
    try:
        lines = build_command_lines(req.model_dump(), state.config)
        return {"ok": True, "commands": lines, "count": len(lines)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{task_id}")
def get_task(task_id: str):
    task = state.tasks.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    return task.snapshot()


@router.get("/{task_id}/logs")
def get_task_logs(task_id: str, tail: int = MAX_LOG_LINES):
    """返回任务整条终端输出日志（logs 目录下 perf|eval_runID_*.log，兼容旧版 run_dir/full.log）。默认只取末尾 MAX_LOG_LINES 行用于终端展示。"""
    task = state.tasks.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    full_log = task.log_path or task.run_dir / "full.log"
    if not full_log.exists():
        return {"task_id": task_id, "lines": [], "total_lines": 0, "truncated": 0}
    try:
        content = full_log.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"读取日志失败: {e}")
    lines = content.splitlines()
    tail = max(1, int(tail))
    if len(lines) > tail:
        truncated = len(lines) - tail
        lines = lines[-tail:]
    else:
        truncated = 0
    return {"task_id": task_id, "lines": lines, "total_lines": len(lines) + truncated, "truncated": truncated}


@router.patch("/{task_id}/threshold")
def update_threshold(task_id: str, req: UpdateThresholdRequest):
    """实时更新 TPOT 阈值并持久化，前端失焦保存。"""
    task = state.tasks.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    task.payload["tpot_threshold_ms"] = float(req.tpot_threshold_ms)
    task.persist()
    task.persist_run_json()
    snap = task.snapshot()
    state.hub.broadcast({"type": "task_updated", "task_id": task_id, "task": snap})
    return snap


@router.post("")
def create_task(req: CreateTaskRequest):
    try:
        task = state.tasks.create_task(req.model_dump())
        return {"ok": True, "task_id": task.task_id, "task": task.snapshot()}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{task_id}/start")
def start_task(task_id: str):
    try:
        task = state.tasks.start_task(task_id)
        return {"ok": True, "task_id": task.task_id, "task": task.snapshot()}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=409, detail=str(e))


@router.post("/{task_id}/stop")
def stop_task(task_id: str):
    task = state.tasks.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    state.tasks.stop_task(task_id)
    return {"ok": True}


@router.delete("/{task_id}")
def delete_task(task_id: str):
    state.tasks.delete_task(task_id)
    return {"ok": True}


@router.post("/{task_id}/export")
def export_task_excel(task_id: str, req: ExportExcelRequest):
    """将前端 Realtime Data 表格当前内容导出为 Excel，写入任务记录缓存目录（run_dir），并返回文件下载。"""
    task = state.tasks.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Realtime"
        if req.headers:
            ws.append(req.headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
        for row in req.rows:
            ws.append(row.get("values", []))
            if row.get("group"):
                # 组标题行：加粗 + 浅蓝底色
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E6F4FF")
        # 写入任务记录缓存文件夹（run_dir，与 run.json / CSV / 日志同目录）
        task.run_dir.mkdir(parents=True, exist_ok=True)
        fname = f"realtime_{task_id}.xlsx"
        path = task.run_dir / fname
        wb.save(path)
        return FileResponse(
            path,
            filename=fname,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {e}")


@router.post("/{task_id}/preview")
def preview_task(req: CreateTaskRequest):
    """预览将要执行的命令。"""
    from benchscope.server.test_manager import build_command_lines
    try:
        lines = build_command_lines(req.model_dump(), state.config)
        return {"ok": True, "commands": lines, "count": len(lines)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
