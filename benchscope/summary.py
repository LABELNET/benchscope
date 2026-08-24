"""日志汇总：CSV 汇总日志与 benchmark-*.xlsx 生成（mean / P99 双面板）。"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

log = logging.getLogger("benchscope.summary")

# xlsx 列定义（与 asserts/benchmark-260821.xlsx 对齐，末尾追加 单用户）
XLSX_HEADERS = [
    "GPU", "模型", "精度", "推理框架", "输入长度", "输出长度", "并发数",
    "output", "peakoutput", "total", "ttft", "itl", "tpot", "单用户",
]


def _fmt(value, digits: int = 2) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return str(value)


def write_summary_csv(
    path: Path,
    rows: Iterable[dict],
    p99: bool = False,
    append: bool = False,
    case_header: bool = False,
    case: dict | None = None,
    meta: dict | None = None,
) -> Path:
    """按用例分组写汇总 CSV（兼容 asserts/logs 旧格式）。

    rows: [{case, label, input_len, output_len, concurrency, metrics}]
    p99=False 时取 mean 指标；p99=True 时取 P99 指标。
    append=True 时为增量追加模式（配合 case_header 控制块头写入）。
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = "_p99" if p99 else ""
    key = "p99" if p99 else "mean"

    def metric(r, name):
        m = r.get("metrics", {})
        return m.get(f"{name}_{key}", m.get(name, ""))

    rows = list(rows)
    if not append:
        with open(path, "w", encoding="utf-8") as f:
            pass  # 清空

    with open(path, "a", encoding="utf-8") as f:
        if append:
            if case_header and case:
                gpu_count = (meta or {}).get("gpu", "") or ""
                f.write("=" * 60 + "\n")
                f.write(
                    f"测试条件：{case.get('label')} | 输入={case.get('input_len')} | "
                    f"输出={case.get('output_len')} | 部署GPU={gpu_count}\n"
                )
                f.write("=" * 60 + "\n")
                f.write("并发数,Output Token,Peak Output Token,Total Token,TTFT,TPOT,ITL\n")
            for r in rows:
                f.write(
                    f"{r.get('concurrency')},{metric(r, 'output')},{metric(r, 'peakoutput')},"
                    f"{metric(r, 'total')},{metric(r, 'ttft')},{metric(r, 'tpot')},{metric(r, 'itl')}\n"
                )
            return path

        # 全量模式：按用例分组
        groups: dict = {}
        for r in rows:
            groups.setdefault(r.get("label", ""), []).append(r)
        for label, items in groups.items():
            first = items[0]
            f.write("=" * 60 + "\n")
            f.write(
                f"测试条件：{label} | 输入={first.get('input_len')} | "
                f"输出={first.get('output_len')} | 部署GPU={(meta or {}).get('gpu', '')}\n"
            )
            f.write("=" * 60 + "\n")
            f.write("并发数,Output Token,Peak Output Token,Total Token,TTFT,TPOT,ITL\n")
            for r in items:
                f.write(
                    f"{r.get('concurrency')},{metric(r, 'output')},{metric(r, 'peakoutput')},"
                    f"{metric(r, 'total')},{metric(r, 'ttft')},{metric(r, 'tpot')},{metric(r, 'itl')}\n"
                )
            f.write("\n")
    return path


def write_xlsx(path: Path, rows: Iterable[dict], meta: dict) -> Path:
    """生成 benchmark-*.xlsx，含 均值 与 P99 两个 sheet。"""
    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    best_fill = PatternFill("solid", fgColor="FFF3CD")

    for sheet_name, key in (("均值 Mean", "mean"), ("P99", "p99")):
        ws = wb.create_sheet(sheet_name)
        ws.append(XLSX_HEADERS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            m = r.get("metrics", {})
            concurrency = r.get("concurrency")
            # 最佳行（tpot 最接近且低于阈值）加亮
            best = r.get("best", False)
            values = [
                meta.get("gpu", ""),
                meta.get("model", ""),
                meta.get("precision", ""),
                meta.get("framework", ""),
                r.get("input_len", ""),
                r.get("output_len", ""),
                concurrency,
                _fmt(m.get(f"output_{key}", m.get("output_mean", m.get("output")))),
                _fmt(m.get(f"peakoutput_{key}", m.get("peakoutput_mean", m.get("peakoutput")))),
                _fmt(m.get(f"total_{key}", m.get("total_mean", m.get("total")))),
                _fmt(m.get(f"ttft_{key}", m.get("ttft"))),
                _fmt(m.get(f"itl_{key}", m.get("itl"))),
                _fmt(m.get(f"tpot_{key}", m.get("tpot"))),
                _fmt(m.get("single_user")),
            ]
            ws.append(values)
            if best:
                for cell in ws[ws.max_row]:
                    cell.fill = best_fill

        # 列宽
        for col, _ in enumerate(XLSX_HEADERS, start=1):
            ws.column_dimensions[chr(64 + col)].width = 14
        ws.freeze_panes = "A2"

    wb.save(path)
    return path
